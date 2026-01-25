"""
Zeiterfassung App - Einfache Version mit Excel-Export
Features: Kunden, Timer, Einträge, Excel-Export (ohne Geldberechnung)
"""
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.scrollview import ScrollView
from kivy.properties import ListProperty, StringProperty, NumericProperty
from kivy.clock import Clock
from kivy.utils import platform
import os
import time
from datetime import datetime
import db


class ZeiterfassungApp(App):
    """Hauptapp mit Timer und Datenbank-Integration"""
    
    def build(self):
        self.root_widget = RootWidget()
        return self.root_widget
    
    def get_db_path(self):
        """Plattform-spezifischer Datenbank-Pfad"""
        if platform == 'android':
            from android.storage import app_storage_path
            return os.path.join(app_storage_path(), 'zeiterfassung.db')
        else:
            return os.path.join(os.path.dirname(__file__), 'zeiterfassung.db')


class RootWidget(BoxLayout):
    """Haupt-UI"""
    customers = ListProperty([])
    selected_customer = StringProperty('— Bitte wählen —')
    selected_customer_id = NumericProperty(0)
    elapsed_time = StringProperty('00:00:00')
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.spacing = 10
        self.padding = 15
        
        # Timer state
        self._timer_start = 0
        self._timer_paused_time = 0
        self._timer_event = None
        
        # Excel-Speicherpfad (Standard: aktuelles Verzeichnis)
        self.excel_export_path = os.path.dirname(__file__) if os.path.dirname(__file__) else os.getcwd()
        
        # Build UI
        Clock.schedule_once(lambda dt: self.build_ui(), 0)
        Clock.schedule_once(lambda dt: self.load_customers(), 0.1)
    
    def get_db_path(self):
        """DB-Pfad von App holen"""
        return App.get_running_app().get_db_path()
    
    def round_to_quarter_hour(self, hours):
        """Runde Stunden auf 0.25 h (15 Minuten) auf"""
        import math
        return math.ceil(hours * 4) / 4.0
    
    def build_ui(self):
        """Erstelle UI-Komponenten"""
        self.clear_widgets()
        
        # === KUNDEN-AUSWAHL ===
        customer_box = BoxLayout(orientation='vertical', size_hint_y=None, height=120, spacing=5)
        customer_box.add_widget(Label(text='Kunde', size_hint_y=None, height=30, bold=True))
        
        self.customer_spinner = Button(
            text=self.selected_customer,
            size_hint_y=None, 
            height=50,
            background_color=(0.3, 0.3, 0.3, 1)
        )
        self.customer_spinner.bind(on_release=self.show_customer_popup)
        customer_box.add_widget(self.customer_spinner)
        
        # Kunden-Verwaltungs-Buttons
        customer_btn_box = BoxLayout(size_hint_y=None, height=40, spacing=5)
        add_customer_btn = Button(text='+ Neuer Kunde', background_color=(0.2, 0.6, 0.2, 1))
        add_customer_btn.bind(on_release=lambda x: self.add_customer_popup())
        customer_btn_box.add_widget(add_customer_btn)
        
        manage_btn = Button(text='⚙ Verwalten', background_color=(0.5, 0.5, 0.2, 1))
        manage_btn.bind(on_release=lambda x: self.open_customer_management())
        customer_btn_box.add_widget(manage_btn)
        customer_box.add_widget(customer_btn_box)
        
        self.add_widget(customer_box)
        
        # === AKTIVITÄT ===
        activity_box = BoxLayout(orientation='vertical', size_hint_y=None, height=90, spacing=5)
        activity_box.add_widget(Label(text='Tätigkeit', size_hint_y=None, height=30, bold=True))
        self.activity_input = TextInput(hint_text='z.B. Mais dreschen', multiline=False, size_hint_y=None, height=50)
        activity_box.add_widget(self.activity_input)
        self.add_widget(activity_box)
        
        # === TIMER ===
        timer_box = BoxLayout(orientation='vertical', size_hint_y=None, height=140, spacing=5)
        timer_box.add_widget(Label(text='Timer', size_hint_y=None, height=30, bold=True))
        
        self.time_label = Label(text=self.elapsed_time, size_hint_y=None, height=50, font_size='32sp')
        timer_box.add_widget(self.time_label)
        
        timer_buttons = BoxLayout(size_hint_y=None, height=50, spacing=5)
        self.start_btn = Button(text='▶ Start', background_color=(0.2, 0.7, 0.2, 1))
        self.pause_btn = Button(text='⏸ Pause', background_color=(0.7, 0.5, 0.2, 1), disabled=True)
        self.stop_btn = Button(text='⏹ Stop', background_color=(0.7, 0.2, 0.2, 1), disabled=True)
        
        self.start_btn.bind(on_release=lambda x: self.start_timer())
        self.pause_btn.bind(on_release=lambda x: self.pause_timer())
        self.stop_btn.bind(on_release=lambda x: self.stop_timer())
        
        timer_buttons.add_widget(self.start_btn)
        timer_buttons.add_widget(self.pause_btn)
        timer_buttons.add_widget(self.stop_btn)
        timer_box.add_widget(timer_buttons)
        
        # Manueller Eintrag-Button
        manual_btn = Button(text='⏱ Stunden manuell nachtragen', size_hint_y=None, height=45, background_color=(0.3, 0.3, 0.5, 1))
        manual_btn.bind(on_release=lambda x: self.add_manual_entry())
        timer_box.add_widget(manual_btn)
        
        self.add_widget(timer_box)
        
        # === EINTRÄGE-LISTE ===
        entries_label = Label(text='Gespeicherte Einträge', size_hint_y=None, height=40, bold=True)
        self.add_widget(entries_label)
        
        scroll = ScrollView()
        self.entries_list = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        self.entries_list.bind(minimum_height=self.entries_list.setter('height'))
        scroll.add_widget(self.entries_list)
        self.add_widget(scroll)
        
        # === EXCEL-EXPORT ===
        excel_box = BoxLayout(orientation='horizontal', size_hint_y=None, height=50, spacing=5)
        excel_btn = Button(text='📊 Excel-Report erstellen', background_color=(0.2, 0.6, 0.3, 1))
        excel_btn.bind(on_release=lambda x: self.export_excel())
        excel_box.add_widget(excel_btn)
        
        path_btn = Button(text='📁', size_hint_x=None, width=60, background_color=(0.4, 0.4, 0.4, 1))
        path_btn.bind(on_release=lambda x: self.change_excel_path())
        excel_box.add_widget(path_btn)
        
        self.add_widget(excel_box)
    
    def load_customers(self):
        """Lade Kunden aus DB"""
        customers = db.get_all_customers(self.get_db_path())
        self.customers = [c[1] for c in customers]  # c[1] = name
        if not self.customers:
            self.customers = ['— Keine Kunden —']
    
    def show_customer_popup(self, instance):
        """Zeige Kunden-Auswahl"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        
        scroll = ScrollView()
        btn_box = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        btn_box.bind(minimum_height=btn_box.setter('height'))
        
        for customer in self.customers:
            if customer == '— Keine Kunden —':
                continue
            btn = Button(text=customer, size_hint_y=None, height=50)
            btn.bind(on_release=lambda x, c=customer: self.select_customer(c, popup))
            btn_box.add_widget(btn)
        
        scroll.add_widget(btn_box)
        content.add_widget(scroll)
        
        popup = Popup(title='Kunde wählen', content=content, size_hint=(0.9, 0.7))
        popup.open()
    
    def select_customer(self, customer_name, popup):
        """Kunde auswählen"""
        self.selected_customer = customer_name
        self.customer_spinner.text = customer_name
        
        # Hole Customer ID
        customer = db.get_customer_by_name(self.get_db_path(), customer_name)
        if customer:
            self.selected_customer_id = customer[0]
        
        popup.dismiss()
        self.refresh_entries()
    
    def add_customer_popup(self):
        """Popup zum Hinzufügen eines neuen Kunden"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        
        content.add_widget(Label(text='Neuer Kunde', size_hint_y=None, height=30, bold=True))
        
        content.add_widget(Label(text='Name:', size_hint_y=None, height=30))
        name_input = TextInput(multiline=False, size_hint_y=None, height=50)
        content.add_widget(name_input)
        
        # Optionale Felder auskommentiert - können später aktiviert werden
        # content.add_widget(Label(text='Adresse:', size_hint_y=None, height=30))
        # address_input = TextInput(multiline=False, size_hint_y=None, height=50)
        # content.add_widget(address_input)
        
        btn_box = BoxLayout(size_hint_y=None, height=50, spacing=10)
        save_btn = Button(text='Speichern', background_color=(0.2, 0.7, 0.2, 1))
        cancel_btn = Button(text='Abbrechen', background_color=(0.7, 0.2, 0.2, 1))
        
        popup = Popup(title='Neuer Kunde', content=content, size_hint=(0.9, 0.5))
        
        def save_customer(instance):
            name = name_input.text.strip()
            if not name:
                name_input.hint_text = 'Bitte Name eingeben!'
                return
            
            # Stundensatz auf 0 setzen (nicht verwendet)
            db.add_customer(self.get_db_path(), name, 0.0, '', '', '')
            popup.dismiss()
            self.load_customers()
            self.show_message('Erfolg', f'Kunde "{name}" gespeichert!')
        
        save_btn.bind(on_release=save_customer)
        cancel_btn.bind(on_release=popup.dismiss)
        
        btn_box.add_widget(save_btn)
        btn_box.add_widget(cancel_btn)
        content.add_widget(btn_box)
        
        popup.open()
    
    def open_customer_management(self):
        """Öffne Kunden-Verwaltung"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        
        scroll = ScrollView()
        customer_list = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        customer_list.bind(minimum_height=customer_list.setter('height'))
        
        customers = db.get_all_customers(self.get_db_path())
        for customer in customers:
            cust_id, name, hourly_rate, address, email, phone = customer
            
            cust_box = BoxLayout(size_hint_y=None, height=50, spacing=5)
            cust_box.add_widget(Label(text=name, halign='left'))
            
            edit_btn = Button(text='✏', size_hint_x=None, width=50, background_color=(0.3, 0.5, 0.7, 1))
            edit_btn.bind(on_release=lambda x, c=customer: self.edit_customer_popup(c, popup))
            cust_box.add_widget(edit_btn)
            
            del_btn = Button(text='🗑', size_hint_x=None, width=50, background_color=(0.7, 0.2, 0.2, 1))
            del_btn.bind(on_release=lambda x, c_id=cust_id, n=name: self.delete_customer_confirm(c_id, n, popup))
            cust_box.add_widget(del_btn)
            
            customer_list.add_widget(cust_box)
        
        scroll.add_widget(customer_list)
        content.add_widget(scroll)
        
        close_btn = Button(text='Schließen', size_hint_y=None, height=50)
        popup = Popup(title='Kunden verwalten', content=content, size_hint=(0.9, 0.8))
        close_btn.bind(on_release=popup.dismiss)
        content.add_widget(close_btn)
        
        popup.open()
    
    def edit_customer_popup(self, customer, parent_popup):
        """Kunde bearbeiten"""
        cust_id, name, hourly_rate, address, email, phone = customer
        
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        content.add_widget(Label(text=f'Kunde bearbeiten: {name}', size_hint_y=None, height=30, bold=True))
        
        content.add_widget(Label(text='Name:', size_hint_y=None, height=30))
        name_input = TextInput(text=name, multiline=False, size_hint_y=None, height=50)
        content.add_widget(name_input)
        
        btn_box = BoxLayout(size_hint_y=None, height=50, spacing=10)
        save_btn = Button(text='Speichern', background_color=(0.2, 0.7, 0.2, 1))
        cancel_btn = Button(text='Abbrechen', background_color=(0.7, 0.2, 0.2, 1))
        
        popup = Popup(title='Kunde bearbeiten', content=content, size_hint=(0.9, 0.5))
        
        def save_changes(instance):
            new_name = name_input.text.strip()
            if not new_name:
                name_input.hint_text = 'Bitte Name eingeben!'
                return
            
            db.update_customer(self.get_db_path(), cust_id, new_name, 0.0, '', '', '')
            popup.dismiss()
            parent_popup.dismiss()
            self.load_customers()
            self.show_message('Erfolg', f'Kunde "{new_name}" aktualisiert!')
        
        save_btn.bind(on_release=save_changes)
        cancel_btn.bind(on_release=popup.dismiss)
        
        btn_box.add_widget(save_btn)
        btn_box.add_widget(cancel_btn)
        content.add_widget(btn_box)
        
        popup.open()
    
    def delete_customer_confirm(self, customer_id, customer_name, parent_popup):
        """Bestätigung vor Löschen"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        content.add_widget(Label(text=f'Kunde "{customer_name}" wirklich löschen?\\n\\nAlle Einträge gehen verloren!'))
        
        btn_box = BoxLayout(size_hint_y=None, height=50, spacing=10)
        yes_btn = Button(text='Ja, löschen', background_color=(0.7, 0.2, 0.2, 1))
        no_btn = Button(text='Abbrechen', background_color=(0.3, 0.3, 0.3, 1))
        
        popup = Popup(title='Löschen bestätigen', content=content, size_hint=(0.8, 0.4))
        
        def do_delete(instance):
            db.delete_customer(self.get_db_path(), customer_id)
            popup.dismiss()
            parent_popup.dismiss()
            self.load_customers()
            self.show_message('Erfolg', f'Kunde "{customer_name}" gelöscht!')
        
        yes_btn.bind(on_release=do_delete)
        no_btn.bind(on_release=popup.dismiss)
        
        btn_box.add_widget(yes_btn)
        btn_box.add_widget(no_btn)
        content.add_widget(btn_box)
        
        popup.open()
    
    def start_timer(self):
        """Timer starten"""
        if not self.selected_customer_id:
            self.show_message('Fehler', 'Bitte zuerst Kunden auswählen!')
            return
        
        if not self.activity_input.text.strip():
            self.show_message('Fehler', 'Bitte Tätigkeit eingeben!')
            return
        
        self._timer_start = time.time() - self._timer_paused_time
        self._timer_event = Clock.schedule_interval(self.update_timer, 0.1)
        
        self.start_btn.disabled = True
        self.pause_btn.disabled = False
        self.stop_btn.disabled = False
    
    def pause_timer(self):
        """Timer pausieren"""
        if self._timer_event:
            self._timer_event.cancel()
            self._timer_event = None
        
        self._timer_paused_time = time.time() - self._timer_start
        
        self.start_btn.disabled = False
        self.start_btn.text = '▶ Fortsetzen'
        self.pause_btn.disabled = True
    
    def update_timer(self, dt):
        """Timer aktualisieren"""
        elapsed = time.time() - self._timer_start
        hours = int(elapsed // 3600)
        minutes = int((elapsed % 3600) // 60)
        seconds = int(elapsed % 60)
        self.elapsed_time = f'{hours:02d}:{minutes:02d}:{seconds:02d}'
        self.time_label.text = self.elapsed_time
    
    def stop_timer(self):
        """Timer stoppen und speichern"""
        if self._timer_event:
            self._timer_event.cancel()
            self._timer_event = None
        
        total_seconds = time.time() - self._timer_start
        duration_hours = total_seconds / 3600.0
        duration_hours = self.round_to_quarter_hour(duration_hours)
        
        # Speichere in DB
        now = datetime.now()
        start_time = (now - datetime.timedelta(seconds=total_seconds)).strftime('%Y-%m-%d %H:%M:%S')
        end_time = now.strftime('%Y-%m-%d %H:%M:%S')
        
        db.add_entry(
            self.get_db_path(),
            self.selected_customer_id,
            self.activity_input.text.strip(),
            start_time,
            end_time,
            duration_hours,
            ''  # Kein Kommentar beim Timer
        )
        
        # Reset
        self._timer_start = 0
        self._timer_paused_time = 0
        self.elapsed_time = '00:00:00'
        self.time_label.text = self.elapsed_time
        self.activity_input.text = ''
        
        self.start_btn.disabled = False
        self.start_btn.text = '▶ Start'
        self.pause_btn.disabled = True
        self.stop_btn.disabled = True
        
        self.refresh_entries()
        self.show_message('Erfolg', f'{duration_hours:.2f} Stunden gespeichert!')
    
    def add_manual_entry(self):
        """Manuellen Eintrag hinzufügen"""
        if not self.selected_customer_id:
            self.show_message('Fehler', 'Bitte zuerst Kunden auswählen!')
            return
        
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        content.add_widget(Label(text='Stunden manuell nachtragen', size_hint_y=None, height=30, bold=True))
        
        content.add_widget(Label(text='Datum (YYYY-MM-DD):', size_hint_y=None, height=30))
        date_input = TextInput(text=datetime.now().strftime('%Y-%m-%d'), multiline=False, size_hint_y=None, height=50)
        content.add_widget(date_input)
        
        content.add_widget(Label(text='Tätigkeit:', size_hint_y=None, height=30))
        activity_input = TextInput(multiline=False, size_hint_y=None, height=50, hint_text='z.B. Mais dreschen')
        content.add_widget(activity_input)
        
        content.add_widget(Label(text='Stunden (z.B. 2.5):', size_hint_y=None, height=30))
        hours_input = TextInput(multiline=False, size_hint_y=None, height=50, input_filter='float', hint_text='Wird auf 0.25 h aufgerundet')
        content.add_widget(hours_input)
        
        content.add_widget(Label(text='Kommentar:', size_hint_y=None, height=30))
        comment_input = TextInput(multiline=False, size_hint_y=None, height=50, hint_text='Optional')
        content.add_widget(comment_input)
        
        btn_box = BoxLayout(size_hint_y=None, height=50, spacing=10)
        save_btn = Button(text='Speichern', background_color=(0.2, 0.7, 0.2, 1))
        cancel_btn = Button(text='Abbrechen', background_color=(0.7, 0.2, 0.2, 1))
        
        popup = Popup(title='Manueller Eintrag', content=content, size_hint=(0.9, 0.8))
        
        def save_manual(instance):
            date_str = date_input.text.strip()
            activity = activity_input.text.strip()
            comment = comment_input.text.strip()
            
            try:
                hours = float(hours_input.text.strip())
            except:
                hours_input.hint_text = 'Ungültige Zahl!'
                return
            
            if not activity:
                activity_input.hint_text = 'Bitte Tätigkeit eingeben!'
                return
            
            # Runde auf 0.25 h auf
            hours = self.round_to_quarter_hour(hours)
            
            # Erstelle Start- und Endzeit
            try:
                start_dt = datetime.strptime(f'{date_str} 08:00', '%Y-%m-%d %H:%M')
                end_dt = start_dt + datetime.timedelta(hours=hours)
                
                start_str = start_dt.strftime('%Y-%m-%d %H:%M:%S')
                end_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
                
                db.add_entry(
                    self.get_db_path(),
                    self.selected_customer_id,
                    activity,
                    start_str,
                    end_str,
                    hours,
                    comment
                )
                
                popup.dismiss()
                self.refresh_entries()
                self.show_message('Erfolg', f'{hours:.2f} Stunden nachgetragen!')
                
            except ValueError:
                date_input.hint_text = 'Ungültiges Datum-Format!'
        
        save_btn.bind(on_release=save_manual)
        cancel_btn.bind(on_release=popup.dismiss)
        
        btn_box.add_widget(save_btn)
        btn_box.add_widget(cancel_btn)
        content.add_widget(btn_box)
        
        popup.open()
    
    def refresh_entries(self):
        """Einträge-Liste aktualisieren"""
        self.entries_list.clear_widgets()
        
        if not self.selected_customer_id:
            return
        
        entries = db.get_entries_by_customer(self.get_db_path(), self.selected_customer_id)
        
        for entry in reversed(entries):  # Neueste zuerst
            entry_id, activity, start, end, duration, comment = entry
            
            date_str = start.split()[0] if start else 'N/A'
            
            entry_box = BoxLayout(size_hint_y=None, height=60, spacing=5, padding=5)
            
            info_text = f'{date_str} | {activity} | {duration:.2f}h'
            if comment:
                info_text += f'\\n{comment}'
            
            entry_label = Label(text=info_text, halign='left', valign='middle', size_hint_x=0.8)
            entry_label.bind(size=entry_label.setter('text_size'))
            entry_box.add_widget(entry_label)
            
            del_btn = Button(text='🗑', size_hint_x=0.2, background_color=(0.7, 0.2, 0.2, 1))
            del_btn.bind(on_release=lambda x, e_id=entry_id: self.delete_entry(e_id))
            entry_box.add_widget(del_btn)
            
            self.entries_list.add_widget(entry_box)
    
    def delete_entry(self, entry_id):
        """Eintrag löschen"""
        db.delete_entry(self.get_db_path(), entry_id)
        self.refresh_entries()
    
    def change_excel_path(self):
        """Popup zum Ändern des Excel-Speicherpfads"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        
        content.add_widget(Label(text='Aktueller Excel-Speicherpfad:', size_hint_y=None, height=30, bold=True))
        current_path_label = Label(text=self.excel_export_path, size_hint_y=None, height=60, halign='left', valign='middle')
        current_path_label.bind(size=current_path_label.setter('text_size'))
        content.add_widget(current_path_label)
        
        content.add_widget(Label(text='Neuer Pfad:', size_hint_y=None, height=30))
        path_input = TextInput(text=self.excel_export_path, multiline=False, size_hint_y=None, height=50)
        content.add_widget(path_input)
        
        info_label = Label(
            text='Beispiel: C:\\\\Users\\\\Bene\\\\Documents\\\\Zeiterfassung\\n\\nOder leer lassen für aktuelles Verzeichnis',
            size_hint_y=None,
            height=80,
            halign='left',
            valign='middle'
        )
        info_label.bind(size=info_label.setter('text_size'))
        content.add_widget(info_label)
        
        btn_box = BoxLayout(size_hint_y=None, height=50, spacing=10)
        save_btn = Button(text='Speichern', background_color=(0.2, 0.7, 0.2, 1))
        cancel_btn = Button(text='Abbrechen', background_color=(0.7, 0.2, 0.2, 1))
        
        popup = Popup(title='Excel-Speicherpfad ändern', content=content, size_hint=(0.9, 0.6))
        
        def save_path(instance):
            new_path = path_input.text.strip()
            if not new_path:
                new_path = os.path.dirname(__file__) if os.path.dirname(__file__) else os.getcwd()
            
            # Prüfe ob Pfad existiert oder erstelle ihn
            try:
                os.makedirs(new_path, exist_ok=True)
                self.excel_export_path = new_path
                popup.dismiss()
                self.show_message('Erfolg', f'Excel-Pfad geändert:\\n{new_path}')
            except Exception as e:
                path_input.hint_text = f'Ungültiger Pfad: {str(e)}'
        
        save_btn.bind(on_release=save_path)
        cancel_btn.bind(on_release=popup.dismiss)
        
        btn_box.add_widget(save_btn)
        btn_box.add_widget(cancel_btn)
        content.add_widget(btn_box)
        
        popup.open()
    
    def export_excel(self):
        """Exportiere Excel-Report (NUR Stunden, KEIN Geld)"""
        if not self.selected_customer_id:
            self.show_message('Fehler', 'Bitte zuerst Kunden auswählen!')
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Hole Daten
            customer = db.get_customer_by_name(self.get_db_path(), self.selected_customer)
            entries = db.get_entries_by_customer(self.get_db_path(), self.selected_customer_id)
            
            if not entries:
                self.show_message('Fehler', 'Keine Einträge vorhanden!')
                return
            
            # Erstelle Excel-Verzeichnis
            if platform == 'android':
                try:
                    from android.storage import primary_external_storage_path
                    excel_dir = os.path.join(primary_external_storage_path(), 'Documents')
                except:
                    excel_dir = App.get_running_app().user_data_dir
            else:
                excel_dir = self.excel_export_path
            
            os.makedirs(excel_dir, exist_ok=True)
            
            # Sichere Dateinamen-Erstellung
            safe_name = self.selected_customer.replace(' ', '_').replace('/', '_')
            filename = f'Zeiterfassung_{safe_name}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
            excel_path = os.path.join(excel_dir, filename)
            
            # Erstelle Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Arbeitsreport'
            
            # Styling
            header_font = Font(bold=True, size=16, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            table_header_font = Font(bold=True, size=12)
            table_header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Header-Bereich
            ws.merge_cells('A1:D1')
            ws['A1'] = f'Arbeitsreport: {self.selected_customer}'
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            ws['A2'] = 'Erstellt:'
            ws['B2'] = datetime.now().strftime('%d.%m.%Y %H:%M')
            ws['A2'].font = Font(bold=True)
            
            # Tabellen-Header (Zeile 4)
            row_start = 4
            headers = ['Datum', 'Tätigkeit', 'Stunden', 'Kommentar']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_start, column=col, value=header)
                cell.font = table_header_font
                cell.fill = table_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Einträge
            total_hours = 0.0
            row_num = row_start + 1
            
            for entry in entries:
                entry_id, activity, start, end, duration, comment = entry
                total_hours += duration
                
                date_str = start.split()[0] if start else 'N/A'
                
                ws.cell(row=row_num, column=1, value=date_str).border = border
                ws.cell(row=row_num, column=2, value=str(activity)).border = border
                ws.cell(row=row_num, column=3, value=f'{duration:.2f}').border = border
                ws.cell(row=row_num, column=4, value=str(comment or '')).border = border
                
                row_num += 1
            
            # Summen-Zeile
            row_num += 1
            ws.cell(row=row_num, column=1, value='GESAMT').font = Font(bold=True, size=12)
            ws.cell(row=row_num, column=2, value='').border = border
            ws.cell(row=row_num, column=3, value=f'{total_hours:.2f}').font = Font(bold=True, size=12)
            ws.cell(row=row_num, column=3).border = border
            ws.cell(row=row_num, column=4, value='').border = border
            
            # Spaltenbreiten
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 35
            
            # Speichern
            wb.save(excel_path)
            
            self.show_message('Erfolg', f'Excel erstellt:\\n\\n{excel_path}\\n\\nGesamtstunden: {total_hours:.2f} h')
            
        except ImportError as e:
            self.show_message('Fehler', f'openpyxl fehlt!\\nBitte installieren: pip install openpyxl\\n\\n{str(e)}')
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.show_message('Fehler', f'Excel-Fehler:\\n{str(e)}\\n\\nDetails in Console')
            print("Excel Export Fehler:", error_detail)
    
    def show_message(self, title, message):
        """Zeige Popup-Nachricht"""
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        content.add_widget(Label(text=message))
        btn = Button(text='OK', size_hint_y=None, height=50)
        content.add_widget(btn)
        
        popup = Popup(title=title, content=content, size_hint=(0.8, 0.4))
        btn.bind(on_release=popup.dismiss)
        popup.open()


if __name__ == '__main__':
    ZeiterfassungApp().run()
