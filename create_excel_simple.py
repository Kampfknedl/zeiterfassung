"""
Test-Script für Excel-Export mit openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

# Testdaten
customer_name = "Test-Kunde"
hourly_rate = 45.50
entries = [
    ("2026-01-20 08:00:00", "Mais dreschen", 3.5, "Feld 12a"),
    ("2026-01-21 09:00:00", "Gülle fahren", 5.0, "Betrieb Schmidt"),
    ("2026-01-22 07:30:00", "Pflügen", 4.25, "Feld 7"),
]

# Excel erstellen
wb = Workbook()
ws = wb.active
ws.title = 'Arbeitsreport'

# Styling
header_font = Font(bold=True, size=16, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
table_header_font = Font(bold=True, size=12)
table_header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Header
ws.merge_cells('A1:E1')
ws['A1'] = f'Arbeitsreport: {customer_name}'
ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25

ws['A2'] = 'Erstellt:'
ws['B2'] = datetime.now().strftime('%d.%m.%Y %H:%M')
ws['A2'].font = Font(bold=True)

ws['A3'] = 'Stundensatz:'
ws['B3'] = f'{hourly_rate:.2f} €/h'
ws['A3'].font = Font(bold=True)

# Tabellen-Header
row_start = 5
headers = ['Datum', 'Tätigkeit', 'Stunden', 'Kommentar', 'Betrag (€)']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=row_start, column=col, value=header)
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Einträge
total_hours = 0.0
total_amount = 0.0
row_num = row_start + 1

for start_time, activity, duration, comment in entries:
    total_hours += duration
    date_str = start_time.split()[0]
    amount = duration * hourly_rate
    total_amount += amount
    
    ws.cell(row=row_num, column=1, value=date_str).border = border
    ws.cell(row=row_num, column=2, value=activity).border = border
    ws.cell(row=row_num, column=3, value=f'{duration:.2f}').border = border
    ws.cell(row=row_num, column=4, value=comment).border = border
    ws.cell(row=row_num, column=5, value=f'{amount:.2f}').border = border
    
    row_num += 1

# Summen
row_num += 1
ws.cell(row=row_num, column=1, value='GESAMT').font = Font(bold=True, size=12)
ws.cell(row=row_num, column=2, value='').border = border
ws.cell(row=row_num, column=3, value=f'{total_hours:.2f}').font = Font(bold=True, size=12)
ws.cell(row=row_num, column=3).border = border
ws.cell(row=row_num, column=4, value='').border = border
ws.cell(row=row_num, column=5, value=f'{total_amount:.2f}').font = Font(bold=True, size=12)
ws.cell(row=row_num, column=5).border = border

# Spaltenbreiten
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 12

# Speichern
filename = f'Test_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
wb.save(filename)

print(f"✅ Excel erstellt: {os.path.abspath(filename)}")
print(f"📊 Total: {total_hours:.2f} h = {total_amount:.2f} €")
